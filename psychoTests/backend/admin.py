from django.contrib import admin
import openpyxl
from django.http import HttpResponse
from .models import TestNSI, TestResult, User

# Регистрируем модель TestResult в админке
@admin.register(TestResult)
class TestResultAdminPanel(admin.ModelAdmin):
    list_display = ('id', 'test_id', 'user_id', 'try_number', 'accuracy', 'number_all_answers', 'number_correct_answers', 'complete_time')  # Поля, которые будут отображаться в списке
    search_fields = ('user_id', 'test_id')  # Поля, по которым можно искать
    list_filter = ('test_id', 'user_id')  # Фильтры в админке




@admin.register(TestNSI)
class TestNSIPanel(admin.ModelAdmin):
    list_display = ('test_id', 'test_name') # Поля, которые будут отображаться в списке

@admin.register(User)
class User(admin.ModelAdmin):
    list_display = [field.name for field in User._meta.get_fields() if not field.many_to_many and not field.one_to_many]
    actions = ['export_as_xlsx']
    def export_as_xlsx(self, request, queryset):
        # Создаем HTTP-ответ с типом контента для Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=users.xlsx'

        # Создаем новую книгу Excel и активный лист
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Users'

        # Заголовки столбцов с информацией о пользователе
        headers = ['ИД', 'Возраст', 'Образование', 'Специальность', 'Место жительства', 'Рост', 'Вес', 'Ведущая рука',
                   'Заболевнаия', 'Курение', 'Алкоголь', 'Спорт', 'Бесонница', 'Текущее самочувствие', 'Геймер']
        for col_idx, header in enumerate(headers, start=1):
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            ws.cell(row=1, column=col_idx, value=header)

        # Заголовки для каждого теста с подзаголовками
        tests_info = TestNSI.objects.all()

        test_fields = ['#', '% выполнения', 'Время выполнения', 'Количество верных ответов пользователя', 'Количество верных ответов в тесте']
        start_tests_cols = len(headers)+1
        test_col_count = len(test_fields)
        for col_idx, test in enumerate(tests_info, start=0):
            ws.cell(row=1, column=start_tests_cols + col_idx * test_col_count, value=test.test_name)
            ws.merge_cells(start_row=1, start_column=start_tests_cols+col_idx*test_col_count, end_row=1, end_column=start_tests_cols+col_idx*test_col_count + test_col_count-1)
            for col,name in enumerate(test_fields, start=0):
                ws.cell(row=2, column=start_tests_cols+col_idx*test_col_count+col, value=name)
        row = 2
        # Заполняем данными
        for user in queryset:
            max_tries = 1
            user_info = [user.id, user.age, user.education, user.speciality, user.residence, user.height, user.weight,user.lead_hand,
            user.diseases, user.smoking, user.alcohol, user.sport, user.insomnia, user.current_health,user.gaming]
            for col_test, test in enumerate(tests_info, start=0):
                user_results = TestResult.objects.filter(test=test.test_id, user=user.id)
                if len(user_results)>max_tries:
                    max_tries = len(user_results)
                for row_try, res in enumerate(user_results, start=1):
                    test_res = [row_try, res.accuracy, res.complete_time, res.number_correct_answers, res.number_all_answers]
                    for j, field in enumerate(test_res, start=0):
                        ws.cell(row=row+row_try, column=start_tests_cols+col_test*test_col_count+j, value=field)
            for col, user_param in enumerate(user_info, start=1):
                ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + max_tries, end_column=col)
                ws.cell(row=row+1, column=col, value=user_param)
            row+=max_tries

        # Сохраняем книгу в ответ
        wb.save(response)
        return response

    export_as_xlsx.short_description = "Export Selected as XLSX"